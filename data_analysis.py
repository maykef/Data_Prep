#!/usr/bin/env python3
"""
Detect real populated rows per sheet in a large Excel workbook.

Strategy:
- A row is "populated" if ANY cell in that row is non-blank.
- Scan bottom-up from the sheet's max_row until we find the last populated row.
- Report both Excel's max_row and the detected real row count.

Safe for large sheets (1M rows).
Parallelised with processes (no threads).
"""

import os
import sys
import time
import logging
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

# -------- CONFIG --------
INPUT_XLSX = "Data_Base.xlsx"
OUTPUT_REPORT = "real_row_counts.md"
HEADER_ROW = 1  # assumes header is first row; adjust if different
MAX_WORKERS = None  # None -> use all CPU cores
# ------------------------

# Prevent oversubscription by BLAS libraries
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("real_row_counter")


def detect_last_populated_row(path: str, sheet: str, header_row: int) -> Dict:
    """Detect last populated row in a sheet (bottom-up scan)."""
    started = time.time()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    last_populated = header_row
    # scan bottom-up
    for r in range(max_row, header_row, -1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if any(v not in (None, "", " ") for v in values):
            last_populated = r
            break

    wb.close()
    elapsed = round(time.time() - started, 3)

    return {
        "sheet": sheet,
        "header_row": header_row,
        "max_row": max_row,
        "last_populated_row": last_populated,
        "real_row_count": max(0, last_populated - header_row),
        "elapsed_sec": elapsed,
    }


def analyze_workbook(path: str) -> Dict[str, Dict]:
    xl = pd.ExcelFile(path, engine="openpyxl")
    sheets = xl.sheet_names

    ctx = (
        mp.get_context("spawn") if sys.platform == "darwin"
        else mp.get_context("forkserver")
    )
    workers = MAX_WORKERS or os.cpu_count() or 4

    results = []
    with ProcessPoolExecutor(max_workers=workers, mp_context=ctx) as ex:
        futs = {ex.submit(detect_last_populated_row, path, s, HEADER_ROW): s for s in sheets}
        for fut in tqdm(as_completed(futs), total=len(futs), desc="Sheets", unit="sheet"):
            results.append(fut.result())

    # keep order
    order = {s: i for i, s in enumerate(sheets)}
    results.sort(key=lambda r: order[r["sheet"]])
    return results


def write_report(results, outfile: str):
    lines = []
    lines.append(f"# Real row count analysis: {INPUT_XLSX}\n")
    for r in results:
        lines.append(f"## Sheet: **{r['sheet']}**")
        lines.append(f"- Header row: **{r['header_row']}**")
        lines.append(f"- Excel max_row: **{r['max_row']}**")
        lines.append(f"- Last populated row: **{r['last_populated_row']}**")
        lines.append(f"- Real row count: **{r['real_row_count']}**")
        lines.append(f"- Elapsed: {r['elapsed_sec']} s")
        lines.append("")

    with open(outfile, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return outfile


if __name__ == "__main__":
    t0 = time.time()
    try:
        results = analyze_workbook(INPUT_XLSX)
        path = write_report(results, OUTPUT_REPORT)

        # print preview
        for r in results:
            logger.info("%s | real rows: %d (last=%d, excel=%d)",
                        r["sheet"], r["real_row_count"], r["last_populated_row"], r["max_row"])
        logger.info("✅ Report written to %s", path)
    finally:
        logger.info("Total runtime: %.2f s", time.time() - t0)
